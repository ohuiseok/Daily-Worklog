"""Lightweight file analyzers that avoid uploadable raw content."""

from __future__ import annotations

import csv
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Literal
from xml.etree import ElementTree

from openpyxl import load_workbook

from worklog.scanner import ScannedFile


AnalysisProgressCallback = Callable[[int, int], None]


TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".json",
    ".yaml",
    ".yml",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".py",
    ".html",
    ".css",
}
CSV_EXTENSIONS = {".csv"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
WORD_EXTENSIONS = {".docx"}


@dataclass(frozen=True)
class TextAnalysis:
    line_count: int
    char_count: int
    title_candidate: str | None = None
    keyword_candidates: list[str] = field(default_factory=list)
    preview_lines: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class CsvAnalysis:
    row_count: int
    column_count: int
    columns: list[str] = field(default_factory=list)
    encoding: str = "utf-8-sig"
    sample_rows: list[list[str]] = field(default_factory=list)


@dataclass(frozen=True)
class ExcelSheetAnalysis:
    name: str
    row_count: int
    column_count: int
    header_candidates: list[str] = field(default_factory=list)
    sample_rows: list[list[str]] = field(default_factory=list)


@dataclass(frozen=True)
class ExcelAnalysis:
    sheets: list[ExcelSheetAnalysis] = field(default_factory=list)


@dataclass(frozen=True)
class FileAnalysis:
    path: str
    kind: Literal["text", "csv", "excel", "unsupported", "error"]
    data: TextAnalysis | CsvAnalysis | ExcelAnalysis | None = None
    error: str | None = None


def analyze_files(
    scanned_files: list[ScannedFile],
    *,
    max_files: int = 500,
    include_content_preview: bool = True,
    text_preview_lines: int = 20,
    table_preview_rows: int = 5,
    progress: AnalysisProgressCallback | None = None,
) -> list[FileAnalysis]:
    analyses: list[FileAnalysis] = []
    limited_files = scanned_files[:max_files] if max_files > 0 else scanned_files
    total = len(limited_files)
    for index, scanned_file in enumerate(limited_files, start=1):
        analyses.append(
            analyze_file(
                scanned_file,
                include_content_preview=include_content_preview,
                text_preview_lines=text_preview_lines,
                table_preview_rows=table_preview_rows,
            )
        )
        if progress is not None and (index == 1 or index == total or index % 100 == 0):
            progress(index, total)
    if max_files > 0 and len(scanned_files) > max_files:
        analyses.append(
            FileAnalysis(
                path=f"{len(scanned_files) - max_files} files",
                kind="unsupported",
                error="analysis_limit_reached",
            )
        )
    return analyses


def analyze_file(
    scanned_file: ScannedFile,
    *,
    include_content_preview: bool = True,
    text_preview_lines: int = 20,
    table_preview_rows: int = 5,
) -> FileAnalysis:
    try:
        extension = scanned_file.extension
        if extension in TEXT_EXTENSIONS:
            return FileAnalysis(
                path=scanned_file.relative_path,
                kind="text",
                data=analyze_text_file(
                    scanned_file.path,
                    preview_lines=text_preview_lines if include_content_preview else 0,
                ),
            )
        if extension in WORD_EXTENSIONS:
            return FileAnalysis(
                path=scanned_file.relative_path,
                kind="text",
                data=analyze_docx_file(
                    scanned_file.path,
                    preview_lines=text_preview_lines if include_content_preview else 0,
                ),
            )
        if extension in CSV_EXTENSIONS:
            return FileAnalysis(
                path=scanned_file.relative_path,
                kind="csv",
                data=analyze_csv_file(
                    scanned_file.path,
                    preview_rows=table_preview_rows if include_content_preview else 0,
                ),
            )
        if extension in EXCEL_EXTENSIONS:
            return FileAnalysis(
                path=scanned_file.relative_path,
                kind="excel",
                data=analyze_excel_file(
                    scanned_file.path,
                    preview_rows=table_preview_rows if include_content_preview else 0,
                ),
            )
        return FileAnalysis(path=scanned_file.relative_path, kind="unsupported")
    except Exception as exc:  # noqa: BLE001 - analyzers must not stop a run.
        return FileAnalysis(
            path=scanned_file.relative_path,
            kind="error",
            error=f"{exc.__class__.__name__}: {exc}",
        )


def analyze_text_file(
    path: Path,
    max_read_bytes: int = 200 * 1024,
    preview_lines: int = 20,
) -> TextAnalysis:
    raw = path.read_bytes()[:max_read_bytes]
    text = _decode_text(raw)
    lines = text.splitlines()
    title = _title_candidate(lines)
    return TextAnalysis(
        line_count=len(lines),
        char_count=len(text),
        title_candidate=title,
        keyword_candidates=_keyword_candidates(text, title),
        preview_lines=_preview_text_lines(lines, preview_lines),
    )


def analyze_docx_file(path: Path, preview_lines: int = 20) -> TextAnalysis:
    text = _extract_docx_text(path)
    lines = text.splitlines()
    title = _title_candidate(lines)
    return TextAnalysis(
        line_count=len(lines),
        char_count=len(text),
        title_candidate=title,
        keyword_candidates=_keyword_candidates(text, title),
        preview_lines=_preview_text_lines(lines, preview_lines),
    )


def analyze_csv_file(path: Path, preview_rows: int = 5) -> CsvAnalysis:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp949"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                sample = file.read(4096)
                file.seek(0)
                dialect = _sniff_csv_dialect(sample)
                reader = csv.reader(file, dialect)
                rows = list(reader)
            if not rows:
                return CsvAnalysis(row_count=0, column_count=0, columns=[], encoding=encoding)
            columns = [str(value).strip() for value in rows[0]]
            return CsvAnalysis(
                row_count=max(len(rows) - 1, 0),
                column_count=len(columns),
                columns=columns,
                encoding=encoding,
                sample_rows=_preview_table_rows(rows[1:], preview_rows),
            )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue

    assert last_error is not None
    raise last_error


def analyze_excel_file(path: Path, preview_rows: int = 5) -> ExcelAnalysis:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[ExcelSheetAnalysis] = []
        for sheet in workbook.worksheets:
            header_candidates: list[str] = []
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            for value in first_row:
                if value is None:
                    continue
                header_candidates.append(str(value).strip())

            sample_rows = []
            if preview_rows > 0:
                sample_rows = [
                    _stringify_row(row)
                    for row in sheet.iter_rows(
                        min_row=2,
                        max_row=1 + preview_rows,
                        values_only=True,
                    )
                    if any(value is not None for value in row)
                ]

            sheets.append(
                ExcelSheetAnalysis(
                    name=sheet.title,
                    row_count=sheet.max_row or 0,
                    column_count=sheet.max_column or 0,
                    header_candidates=header_candidates,
                    sample_rows=sample_rows,
                )
            )
        return ExcelAnalysis(sheets=sheets)
    finally:
        workbook.close()


def analysis_summary(analyses: list[FileAnalysis]) -> dict[str, int]:
    summary = {"text": 0, "csv": 0, "excel": 0, "unsupported": 0, "error": 0}
    for analysis in analyses:
        summary[analysis.kind] += 1
    return summary


def analysis_to_safe_dict(analysis: FileAnalysis) -> dict[str, Any]:
    """Return a dict suitable for dry-run output without raw content."""
    base: dict[str, Any] = {"path": analysis.path, "kind": analysis.kind}
    if analysis.error:
        base["error"] = analysis.error

    data = analysis.data
    if isinstance(data, TextAnalysis):
        base["data"] = {
            "line_count": data.line_count,
            "char_count": data.char_count,
            "title_candidate": data.title_candidate,
            "keyword_candidates": data.keyword_candidates,
            "preview_lines": data.preview_lines,
        }
    elif isinstance(data, CsvAnalysis):
        base["data"] = {
            "row_count": data.row_count,
            "column_count": data.column_count,
            "columns": data.columns,
            "encoding": data.encoding,
            "sample_rows": data.sample_rows,
        }
    elif isinstance(data, ExcelAnalysis):
        base["data"] = {
            "sheets": [
                {
                    "name": sheet.name,
                    "row_count": sheet.row_count,
                    "column_count": sheet.column_count,
                    "header_candidates": sheet.header_candidates,
                    "sample_rows": sheet.sample_rows,
                }
                for sheet in data.sheets
            ]
        }
    return base


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        document_xml = archive.read("word/document.xml")

    root = ElementTree.fromstring(document_xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        parts = [
            text_node.text or ""
            for text_node in paragraph.findall(".//w:t", namespace)
        ]
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _preview_text_lines(lines: list[str], max_lines: int) -> list[str]:
    if max_lines <= 0:
        return []
    preview: list[str] = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        preview.append(stripped[:240])
        if len(preview) >= max_lines:
            break
    return preview


def _preview_table_rows(rows: list[list[str]], max_rows: int) -> list[list[str]]:
    if max_rows <= 0:
        return []
    return [_stringify_row(row) for row in rows[:max_rows]]


def _stringify_row(row: Any) -> list[str]:
    return ["" if value is None else str(value).strip()[:120] for value in row]


def _title_candidate(lines: list[str]) -> str | None:
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("#"):
            stripped = stripped.lstrip("#").strip()
            return stripped[:120] or None
        return None
    return None


def _sniff_csv_dialect(sample: str) -> csv.Dialect:
    if not sample.strip():
        return csv.excel
    try:
        dialect = csv.Sniffer().sniff(sample)
        if (
            not isinstance(dialect.delimiter, str)
            or len(dialect.delimiter) != 1
            or dialect.delimiter not in {",", ";", "\t", "|"}
        ):
            return csv.excel
        return dialect
    except csv.Error:
        return csv.excel


def _keyword_candidates(text: str, title: str | None) -> list[str]:
    words: dict[str, int] = {}
    source = f"{title or ''}\n{text}"
    for raw_word in source.replace("_", " ").replace("-", " ").split():
        word = raw_word.strip(".,:;()[]{}<>\"'`").lower()
        if len(word) < 3:
            continue
        if word in {"the", "and", "for", "with", "from", "this", "that"}:
            continue
        words[word] = words.get(word, 0) + 1

    return [
        word
        for word, _count in sorted(words.items(), key=lambda item: (-item[1], item[0]))[
            :10
        ]
    ]
